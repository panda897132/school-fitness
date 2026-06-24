"""Excel 导入/导出"""

import os
import re
import datetime
import logging
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from config import CN_TO_NUM, GRADE_ITEMS, NUM_TO_CN

logger = logging.getLogger(__name__)


def _normalize_header(hdr):
    """规范化表头：全角数字/符号转半角，去除不可见字符"""
    if not hdr:
        return hdr
    result = []
    for ch in hdr:
        code = ord(ch)
        if 0xFF10 <= code <= 0xFF19:
            result.append(chr(code - 0xFF10 + ord('0')))
        elif 0xFF21 <= code <= 0xFF3A:
            result.append(chr(code - 0xFF21 + ord('A')))
        elif 0xFF41 <= code <= 0xFF5A:
            result.append(chr(code - 0xFF41 + ord('a')))
        elif code == 0xFF0A:
            result.append('*')
        elif code == 0xFF0D:
            result.append('-')
        elif code == 0xFF08:
            result.append('(')
        elif code == 0xFF09:
            result.append(')')
        elif code == 0xFF1A:
            result.append(':')
        elif code == 0xFF0F:
            result.append('/')
        elif code == 0x3000:
            result.append(' ')
        else:
            result.append(ch)
    return ''.join(result)


def _is_valid_value(val):
    """检查是否为有效值（排除错误值和空值）"""
    if val is None:
        return False
    s = str(val).strip()
    if s == '' or s.startswith('#') or s == 'None':
        return False
    return True


def _parse_run_time(raw):
    """解析跑步时间值
    
    支持三种格式:
    - float (如 1.43 表示 1分43秒)
    - str 时间格式 (如 "1'43")
    - 纯秒数
    """
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        v = float(raw)
        # 检测分钟.秒格式：1.0~10.0 区间，小数部分可解释为秒(0-59)
        if 0.5 <= v < 10.0:
            minutes = int(v)
            seconds_part = round((v - minutes) * 100)
            if 0 <= seconds_part < 60:
                return minutes * 60 + seconds_part
        return v
    # 字符串格式
    raw_str = str(raw).strip()
    m = re.match(r"(\d+)'(\d+)", raw_str)
    if m:
        return float(m.group(1)) * 60 + float(m.group(2))
    try:
        return float(raw_str)
    except (ValueError, TypeError):
        return 0


def _parse_number(raw):
    """通用数值解析：兼容 float/int/str/datetime.time 等 Excel 单元格类型"""
    if raw is None:
        return None
    if isinstance(raw, (int, float)):
        return float(raw)
    if isinstance(raw, datetime.time):
        return raw.hour * 3600.0 + raw.minute * 60.0 + raw.second + raw.microsecond / 1000000.0
    if isinstance(raw, datetime.datetime):
        return raw.hour * 3600.0 + raw.minute * 60.0 + raw.second + raw.microsecond / 1000000.0
    raw_str = str(raw).strip()
    if raw_str == '':
        return None
    try:
        return float(raw_str)
    except (ValueError, TypeError):
        pass
    m = re.match(r'^(\d+):(\d+(?:\.\d+)?)$', raw_str)
    if m:
        return float(m.group(1)) * 60 + float(m.group(2))
    m = re.match(r"(\d+)'(\d+(?:\.\d+)?)$", raw_str)
    if m:
        return float(m.group(1)) * 60 + float(m.group(2))
    return None


def parse_filename_for_import(filename):
    """从文件名智能提取年级和班级编号
    
    支持格式:
      '6.1学生体质健康测试.xlsx' → {'grade': 6, 'class_id': '601', 'class_name': '六(01)班'}
      '501班.xlsx' → {'grade': 5, 'class_id': '501', 'class_name': '五(01)班'}
      '五年级(01)班.xlsx' → {'grade': 5, 'class_id': '501', 'class_name': '五(01)班'}
      '1.xlsx' → {'grade': 1, 'class_id': None, 'class_name': None}
    """
    
    name = os.path.splitext(os.path.basename(filename))[0]
    result = {'grade': None, 'class_id': None, 'class_name': None}
    
    # 规则1: "{grade}.{class_num}" 开头
    m = re.match(r'^(\d)\.(\d{1,2})', name)
    if m:
        g, c = int(m.group(1)), int(m.group(2))
        if 1 <= g <= 6:
            cn = NUM_TO_CN.get(g, str(g))
            result['grade'] = g
            result['class_id'] = f'{g}{c:02d}'
            result['class_name'] = f'{cn}({c:02d})班'
            return result
    
    # 规则2: 3位班级编号开头
    m = re.match(r'^(\d{3})', name)
    if m:
        cid = m.group(1)
        g = int(cid[0])
        if 1 <= g <= 6:
            cn = NUM_TO_CN.get(g, str(g))
            result['grade'] = g
            result['class_id'] = cid
            result['class_name'] = f'{cn}({cid[-2:]})班'
            return result
    
    # 规则3: 中文 "N年级(M)班"
    m = re.match(r'(.)年级\(?(\d+)\)?班', name)
    if m:
        g = CN_TO_NUM.get(m.group(1), 1)
        if 1 <= g <= 6:
            c = int(m.group(2))
            result['grade'] = g
            result['class_id'] = f'{g}{c:02d}'
            result['class_name'] = f'{m.group(1)}({c:02d})班'
            return result
    
    # 规则4: 单个数字开头（仅年级）
    m = re.match(r'^(\d)', name)
    if m:
        g = int(m.group(1))
        if 1 <= g <= 6:
            result['grade'] = g
            return result
    
    return result


def quick_scan_excel(filepath):
    """快速扫描Excel文件：检测格式、统计班级数和行数
    
    Returns:
        {'format': 'old'|'new'|'unknown',
         'sheets': [...],
         'total_rows': N,
         'classes': {'101': {'grade': 1, 'rows': 40}, ...}}
    """
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    except Exception:
        return {'format': 'unknown', 'sheets': [], 'total_rows': 0, 'classes': {}}
    
    grade_names = {'一年级': 1, '二年级': 2, '三年级': 3, '四年级': 4, '五年级': 5, '六年级': 6}
    has_grade_sheets = any(sn in grade_names for sn in wb.sheetnames)
    
    result = {
        'format': 'old' if has_grade_sheets else 'new',
        'sheets': list(wb.sheetnames),
        'total_rows': 0,
        'classes': {}
    }
    
    if has_grade_sheets:
        # 旧格式: 扫描每个年级Sheet的A列获取班级编号
        for sn in wb.sheetnames:
            if sn in grade_names:
                ws = wb[sn]
                grade = grade_names[sn]
                class_ids = set()
                row_count = 0
                for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
                    v = row[0]
                    if v is not None and str(v).strip():
                        cid = str(int(float(str(v)))) if '.' in str(v) else str(v).strip()
                        if len(cid) == 2:
                            cid = f'{grade}{cid}'
                        class_ids.add(cid)
                        row_count += 1
                for cid in sorted(class_ids):
                    result['classes'][cid] = {'grade': grade, 'rows': row_count}
                result['total_rows'] += row_count
    else:
        # 新格式: 尝试检测表头
        for sn in wb.sheetnames:
            if '年级' in sn:
                continue
            ws = wb[sn]
            # 找表头行
            header_row = None
            for r in range(1, min(ws.max_row + 1, 5)):
                for c in range(1, min(ws.max_column + 1, 30)):
                    val = str(ws.cell(row=r, column=c).value or '')
                    if '姓名' in val:
                        header_row = r
                        break
                if header_row:
                    break
            
            if header_row:
                row_count = 0
                for row in ws.iter_rows(min_row=header_row + 1, max_col=1, values_only=True):
                    v = row[0]
                    if v is not None and str(v).strip() and '占比' not in str(v):
                        row_count += 1
                result['total_rows'] += row_count
                # 新格式默认一个班级
                if row_count > 0:
                    result['classes']['_new_'] = {'grade': 0, 'rows': row_count}
    
    wb.close()
    return result


def import_from_excel(filepath, grade_hint=None, class_prefix=None):
    """从Excel导入学生数据
    
    兼容两种格式：
    1. 旧模板格式：每个年级一个工作表（一年级/二年级/.../六年级）
    2. 新格式：单工作表"学生成绩"，表头含"总成绩"
    
    参数:
        grade_hint: 年级提示（新格式下优先使用，否则自动推断）
        class_prefix: 班级编号前缀（新格式下使用，否则自动生成）
    
    返回:
        {
            "success": True/False,
            "message": "描述",
            "data": {"classes": {"101": {"grade": 1, "name": "一(1)班", "students": [...]}}}
        }
    """
    try:
        # NOTE: 未使用 read_only=True。旧/新格式导入均大量使用 ws.cell(r,c).value
        # 随机访问模式，read_only 下不支持，必须改用 iter_rows() 遍历。
        # 改写工作量较大（~60+处 .cell() 调用需逐一改为迭代器索引映射）。
        # 格式快速检测（main_window.py:448）已使用 read_only=True 降低内存峰值。
        # 后续如有大文件需求，可优先将 _import_new_format 改为 iter_rows 方式。
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except Exception as e:
        return {"success": False, "message": "无法打开文件，请检查文件格式或是否被其他程序占用", "data": None}
    
    grade_sheet_names = {
        '一年级': 1, '二年级': 2, '三年级': 3,
        '四年级': 4, '五年级': 5, '六年级': 6
    }
    
    # 检测是否有旧格式的年级工作表
    has_old_format = any(sn in grade_sheet_names for sn in wb.sheetnames)
    
    if has_old_format:
        return _import_old_format(wb, grade_sheet_names)
    
    # 尝试新格式：检测不含"年级"且含"总成绩"的工作表
    for sheet_name in wb.sheetnames:
        if '年级' in sheet_name:
            continue
        ws = wb[sheet_name]
        header_row = _detect_new_format_header(ws)
        if header_row is not None:
            result = _import_new_format(ws, header_row, grade_hint, class_prefix)
            wb.close()
            return result
    
    wb.close()
    return {"success": False, "message": "未识别的文件格式，请使用模板或新格式Excel", "data": None}


def _detect_new_format_header(ws):
    """检测新格式表头行（含'姓名'和'总成绩'/'总分'等），返回行号或None"""
    total_keywords = ('总成绩', '总分', '总评', '总得分', '综合得分', '总计')
    for row in range(1, min(ws.max_row + 1, 5)):
        has_name = False
        has_total_score = False
        for col in range(1, ws.max_column + 1):
            val = _normalize_header(str(ws.cell(row=row, column=col).value or '').strip())
            if '姓名' in val:
                has_name = True
            if any(kw in val for kw in total_keywords):
                has_total_score = True
        if has_name and has_total_score:
            return row
    return None


def _import_new_format(ws, header_row, grade_hint, class_prefix):
    """从新格式工作表导入数据
    
    新格式列结构: 姓名|性别|身高|体重|BMI|成绩|肺活量|成绩|50米跑|成绩|...
    每个测试项目后面跟一个"成绩"列(评分)，我们跳过评分列。
    """
    
    # 建立列索引映射
    col_map = {}
    detected_items = set()
    
    for col in range(1, ws.max_column + 1):
        hdr = _normalize_header(str(ws.cell(row=header_row, column=col).value or '').strip())
        if hdr == 'BMI':
            continue
        if '成绩' in hdr or '得分' in hdr or '分数' in hdr or any(kw in hdr for kw in ('总分', '总评', '总计')):
            continue
        if '姓名' in hdr:
            col_map['name'] = col
        elif '性别' in hdr:
            col_map['gender'] = col
        elif hdr == '身高':
            col_map['height'] = col
        elif hdr == '体重':
            col_map['weight'] = col
        elif '肺活量' in hdr:
            col_map['肺活量'] = col
            detected_items.add('肺活量')
        elif ((('50米' in hdr or '50m' in hdr.lower() or '50M' in hdr) and '×8' not in hdr and '*8' not in hdr and '50×8' not in hdr and '50*8' not in hdr
              and '50x8' not in hdr and '50X8' not in hdr and '往返' not in hdr and '折返' not in hdr)):
            col_map['50米跑'] = col
            detected_items.add('50米跑')
        elif '50×8' in hdr or '50*8' in hdr or '50x8' in hdr or '50X8' in hdr or '往返跑' in hdr or '×8' in hdr or '*8' in hdr:
            col_map['50*8折返跑'] = col
            detected_items.add('50*8折返跑')
        elif '坐位' in hdr:
            col_map['坐位体前屈'] = col
            detected_items.add('坐位体前屈')
        elif '跳绳' in hdr:
            col_map['一分钟跳绳'] = col
            detected_items.add('一分钟跳绳')
        elif '仰卧' in hdr:
            col_map['仰卧起坐'] = col
            detected_items.add('仰卧起坐')
    
    # 推断年级
    if grade_hint:
        grade = grade_hint
    elif '50*8折返跑' in detected_items:
        grade = 5
    elif '仰卧起坐' in detected_items:
        grade = 3
    else:
        grade = 1
    
    # 自动生成班级编号
    if class_prefix:
        cid_prefix = str(class_prefix).strip()
    else:
        cid_prefix = f"{grade}01"
    
    # 班级名称
    cn_grade = NUM_TO_CN.get(grade, str(grade))
    class_suffix = cid_prefix[-2:] if len(cid_prefix) >= 2 else cid_prefix
    class_name = f"{cn_grade}({class_suffix})班"
    
    # 读取数据
    students = []
    student_num = 0
    
    for row in range(header_row + 1, ws.max_row + 1):
        name_cell = ws.cell(row=row, column=col_map.get('name', 0)).value
        if not _is_valid_value(name_cell):
            continue
        
        student_num += 1
        student_number = str(student_num).zfill(2)
        
        # 性别映射: "1"→"男", "2"→"女"
        gender_val = str(ws.cell(row=row, column=col_map.get('gender', 0)).value or '').strip()
        if gender_val == '1' or '男' in gender_val:
            gender = '男'
        elif gender_val == '2' or '女' in gender_val:
            gender = '女'
        else:
            gender = '男'
        
        # 身高
        height = None
        if 'height' in col_map:
            hv = ws.cell(row=row, column=col_map['height']).value
            if _is_valid_value(hv):
                height = _parse_number(hv)
        
        # 体重
        weight = None
        if 'weight' in col_map:
            wv = ws.cell(row=row, column=col_map['weight']).value
            if _is_valid_value(wv):
                weight = _parse_number(wv)
        
        # 测试项目值
        tests = {}
        for item_name in GRADE_ITEMS.get(grade, []):
            col_idx = col_map.get(item_name)
            if col_idx:
                raw = ws.cell(row=row, column=col_idx).value
                if _is_valid_value(raw):
                    parsed = _parse_number(raw)
                    if parsed is not None and item_name == '50*8折返跑' and 0 < parsed < 60:
                        minutes = int(parsed)
                        secs = round((parsed - minutes) * 100)
                        if 0 <= secs < 60:
                            parsed = minutes * 60 + secs
                    if parsed is not None:
                        tests[item_name] = parsed
                    else:
                        tests[item_name] = None
                else:
                    tests[item_name] = None
            else:
                tests[item_name] = None
        
        student = {
            'id': f"{cid_prefix}{student_number}",
            'name': str(name_cell).strip(),
            'student_number': student_number,
            'student_code': '',
            'gender': gender,
            'height': height,
            'weight': weight,
            'bmi': None,
            'tests': tests,
            'scores': {},
            'total_score': 0,
            'total_grade': ''
        }
        # 跳过非学生行（如模板中的"单项占比率""总成绩占比率"等汇总行）
        if '占比' in str(name_cell):
            continue
        students.append(student)
    
    return {
        "success": True,
        "message": f"导入完成：成功 {len(students)} 条",
        "data": {
            "classes": {
                cid_prefix: {
                    "grade": grade,
                    "name": class_name,
                    "students": students
                }
            }
        }
    }


def _import_old_format(wb, grade_sheet_names):
    """从旧模板格式导入数据"""
    all_classes = {}
    total_imported = 0
    total_skipped = 0
    
    for sheet_name in wb.sheetnames:
        if sheet_name not in grade_sheet_names:
            continue
        
        grade = grade_sheet_names[sheet_name]
        ws = wb[sheet_name]
        
        # 查找表头行 (包含 "姓名" 的行)
        header_row = None
        for row in range(1, min(ws.max_row + 1, 10)):
            row_values = [str(ws.cell(row=row, column=c).value or '') for c in range(1, ws.max_column + 1)]
            has_name = any('姓名' in v for v in row_values)
            has_number = any('学号' in v or '班级' in v or '学籍' in v for v in row_values)
            if has_name and has_number:
                header_row = row
                break
        
        if header_row is None:
            continue
        
        # 建立列索引映射
        col_map = {}
        for col in range(1, ws.max_column + 1):
            hdr = _normalize_header(str(ws.cell(row=header_row, column=col).value or '').strip())
            if '班级' in hdr:
                col_map['class_id'] = col
            elif '学号' in hdr:
                col_map['student_number'] = col
            elif '姓名' in hdr:
                col_map['name'] = col
            elif '学籍' in hdr:
                col_map['student_code'] = col
            elif '性别' in hdr:
                col_map['gender'] = col
            elif '身高' in hdr:
                col_map['height'] = col
            elif '体重' in hdr:
                col_map['weight'] = col
            elif '肺活量' in hdr:
                col_map['肺活量'] = col
            elif '50米' in hdr and '50*8' not in hdr and '50×8' not in hdr and '50x8' not in hdr and '50X8' not in hdr:
                col_map['50米跑'] = col
            elif '50*8' in hdr or '50×8' in hdr or '50x8' in hdr or '折返' in hdr:
                col_map['50*8折返跑'] = col
            elif '坐位' in hdr:
                col_map['坐位体前屈'] = col
            elif '跳绳' in hdr:
                col_map['一分钟跳绳'] = col
            elif '仰卧' in hdr:
                col_map['仰卧起坐'] = col
        
        # 兼容：可能列名为BMI格式
        if 'BMI(15%)' in col_map:
            col_map['bmi_col'] = col_map.pop('BMI(15%)')
        
        # 读取数据
        students_by_class = {}
        
        for row in range(header_row + 1, ws.max_row + 1):
            name_cell = ws.cell(row=row, column=col_map.get('name', 3)).value
            if not _is_valid_value(name_cell):
                continue  # 空行或错误行跳过
            
            try:
                # 班级编号
                if 'class_id' in col_map:
                    class_id = str(ws.cell(row=row, column=col_map['class_id']).value or '').strip()
                    if '.' in class_id:
                        class_id = str(int(float(class_id)))
                else:
                    class_id = f"{grade}01"
                
                # 如果class_id只有2位，补上年级前缀
                if len(class_id) == 2:
                    class_id = f"{grade}{class_id}"
                
                # 学号
                student_number = ''
                if 'student_number' in col_map:
                    sv = ws.cell(row=row, column=col_map['student_number']).value
                    if sv is not None:
                        student_number = str(int(float(str(sv)))) if '.' in str(sv) else str(sv)
                
                # 姓名
                name = str(name_cell).strip()
                
                # 跳过非学生行（模板中的"单项占比率""总成绩占比率"等）
                if '占比' in name:
                    continue
                
                # 学籍号
                student_code = ''
                if 'student_code' in col_map:
                    sc = ws.cell(row=row, column=col_map['student_code']).value
                    student_code = str(sc).strip() if sc else ''
                
                # 性别
                gender_val = str(ws.cell(row=row, column=col_map.get('gender', 5)).value or '').strip()
                gender = '男' if '男' in gender_val else '女'
                
                # 身高
                height = None
                if 'height' in col_map:
                    hv = ws.cell(row=row, column=col_map['height']).value
                    if _is_valid_value(hv):
                        height = _parse_number(hv)
                
                # 体重
                weight = None
                if 'weight' in col_map:
                    wv = ws.cell(row=row, column=col_map['weight']).value
                    if _is_valid_value(wv):
                        weight = _parse_number(wv)
                
                # 测试项目值
                tests = {}
                for item_name in GRADE_ITEMS.get(grade, []):
                    col_idx = col_map.get(item_name)
                    if col_idx:
                        raw = ws.cell(row=row, column=col_idx).value
                        if _is_valid_value(raw):
                            parsed = _parse_number(raw)
                            if parsed is not None and item_name == '50*8折返跑' and 0 < parsed < 60:
                                minutes = int(parsed)
                                secs = round((parsed - minutes) * 100)
                                if 0 <= secs < 60:
                                    parsed = minutes * 60 + secs
                            if parsed is not None:
                                tests[item_name] = parsed
                            else:
                                tests[item_name] = None
                        else:
                            tests[item_name] = None
                    else:
                        tests[item_name] = None
                
                student = {
                    'id': f"{class_id}{student_number}",
                    'name': name,
                    'student_number': student_number,
                    'student_code': student_code,
                    'gender': gender,
                    'height': height,
                    'weight': weight,
                    'bmi': None,
                    'tests': tests,
                    'scores': {},
                    'total_score': 0,
                    'total_grade': ''
                }
                
                if class_id not in students_by_class:
                    students_by_class[class_id] = {
                        'grade': grade,
                        'name': f"{grade}年级{class_id[-2:]}班",
                        'students': []
                    }
                
                students_by_class[class_id]['students'].append(student)
                total_imported += 1
            except Exception:
                total_skipped += 1
                continue
        
        all_classes.update(students_by_class)
    
    wb.close()
    
    return {
        "success": True,
        "message": f"导入完成：成功 {total_imported} 条，跳过 {total_skipped} 条",
        "data": {"classes": all_classes}
    }


def _to_time_str(val):
    """将秒数或 m.ss 格式转为 m.ss 小数格式（与导入格式一致）

    例: 103秒 → 1.43（1分43秒）
         1.43 → 1.43（m.ss格式原样保留）
    """
    if val is None or val == '':
        return ''
    try:
        v = float(val)
    except (ValueError, TypeError):
        return str(val)
    if v <= 0:
        return ''
    if 0 < v < 60:
        m = int(v)
        s = round((v - m) * 100)
        if 0 <= s < 60:
            return f"{m}.{s:02d}"
    m = int(v // 60)
    s = int(v % 60)
    return f"{m}.{s:02d}"


def export_statistics_report(dm, output_path, scope='全校', grade=None, class_id=None):
    """导出统计分析报告（三层分析版）
    
    Args:
        dm: DataManager 实例
        output_path: 输出文件路径
        scope: '全校'/'年级'/'班级'
        grade: 年级数字 (scope='年级'时)
        class_id: 班级编号 (scope='班级'时)
    """
    try:
        from analysis import analyze_class, analyze_grade, analyze_school
        from config import GRADE_ITEMS
        
        wb = openpyxl.Workbook()
        
        # 样式
        title_font = Font(name='微软雅黑', size=14, bold=True)
        subtitle_font = Font(name='微软雅黑', size=11, bold=True)
        header_font = Font(name='微软雅黑', size=10, bold=True)
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font_white = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')
        green_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))
        center_align = Alignment(horizontal='center', vertical='center')
        
        def _style_header(ws, row, cols, values):
            for ci, v in enumerate(values, 1):
                cell = ws.cell(row=row, column=ci, value=v)
                cell.font = header_font_white
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = border
        
        def _style_row(ws, row, cols, values, bold_first=False):
            for ci, v in enumerate(values, 1):
                cell = ws.cell(row=row, column=ci, value=v if v is not None else '')
                cell.border = border
                cell.alignment = center_align
                if bold_first and ci == 1:
                    cell.font = header_font
        
        # ========== Sheet 1: 综合分析 ==========
        ws_analysis = wb.active
        ws_analysis.title = '综合分析'
        
        title_map = {'全校': '全校', '年级': f'{grade}年级', '班级': f'{class_id}班' if class_id else ''}
        
        # 根据 scope 获取分析数据
        analysis_data = None
        if scope == '班级' and class_id:
            analysis_data = analyze_class(dm, class_id)
        elif scope == '年级' and grade:
            analysis_data = analyze_grade(dm, grade)
        else:
            analysis_data = analyze_school(dm)
        
        if analysis_data and analysis_data.get('total', 0) > 0:
            r = 1
            ws_analysis.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
            ws_analysis.cell(row=r, column=1, value=f'诸葛镇中心小学 — 体质健康分析报告 ({title_map[scope]})').font = title_font
            r += 2
            
            # 基本信息
            ws_analysis.cell(row=r, column=1, value='基本信息').font = subtitle_font
            r += 1
            _style_header(ws_analysis, r, 10, ['指标', '数值'])
            r += 1
            info_rows = [
                ('总人数', str(analysis_data['total'])),
                ('男生', str(analysis_data.get('male_count', '?'))),
                ('女生', str(analysis_data.get('female_count', '?'))),
                ('有效数据', f"{analysis_data['valid_count']} 人"),
                ('数据不全', f"{analysis_data.get('invalid_count', 0)} 人"),
                ('有效数据占比', f"{analysis_data.get('validity_rate', 0)}%"),
            ]
            for label, val in info_rows:
                _style_row(ws_analysis, r, 2, [label, val], bold_first=True)
                r += 1
            
            r += 1
            
            # 总分等级
            tg = analysis_data.get('total_grade', {})
            if tg and tg.get('effective_total', 0) > 0:
                ws_analysis.cell(row=r, column=1, value='总分等级分布').font = subtitle_font
                r += 1
                _style_header(ws_analysis, r, 10, ['等级', '人数', '占比'])
                r += 1
                for grade_label in ['优秀', '良好', '及格', '不及格']:
                    count = tg['distribution'].get(grade_label, 0)
                    pct = f"{round(count / tg['effective_total'] * 100, 1)}%" if tg['effective_total'] > 0 else '0%'
                    _style_row(ws_analysis, r, 3, [grade_label, str(count), pct])
                    r += 1
                excellence_count = sum(tg['distribution'].get(k, 0) for k in ('优秀', '良好'))
                _style_row(ws_analysis, r, 3, ['🏆 优良率',
                    str(excellence_count),
                    f"{tg.get('excellence_rate', 0)}%"], bold_first=True)
                r += 2
            
            # 各项目分析
            items = analysis_data.get('items', [])
            if items:
                # 分离 BMI 和其他项目
                bmi_items = [it for it in items if it['item_name'] == 'BMI']
                other_items = [it for it in items if it['item_name'] != 'BMI']
                
                # ---- BMI 分析区块 ----
                if bmi_items:
                    ws_analysis.cell(row=r, column=1, value='BMI 专项分析').font = subtitle_font
                    r += 1
                    _style_header(ws_analysis, r, 10,
                        ['项目', '正常', '超重', '低体重', '肥胖', '有效总数', '肥胖率'])
                    r += 1
                    for item in bmi_items:
                        dist = item['distribution']
                        eff = item['effective_total']
                        _style_row(ws_analysis, r, 7, [
                            item['item_name'],
                            str(dist.get('正常', 0)),
                            str(dist.get('超重', 0)),
                            str(dist.get('低体重', 0)),
                            str(dist.get('肥胖', 0)),
                            str(eff),
                            f"{item.get('obesity_rate', 0)}%"
                        ], bold_first=True)
                        r += 1
                    r += 1

                # ---- 跳绳附加分分析区块 ----
                jrb = analysis_data.get('jump_rope_bonus')
                if jrb and jrb.get('effective_total', 0) > 0:
                    ws_analysis.cell(row=r, column=1, value='跳绳附加分分析').font = subtitle_font
                    r += 1
                    _style_header(ws_analysis, r, 10,
                        ['指标', '数值'])
                    r += 1
                    _style_row(ws_analysis, r, 2, ['有效人数', str(jrb['effective_total'])], bold_first=True)
                    r += 1
                    _style_row(ws_analysis, r, 2, ['平均附加分', f"{jrb['avg_bonus']} 分"], bold_first=True)
                    r += 1
                    _style_row(ws_analysis, r, 2, ['最高附加分', f"{jrb['max_bonus']} 分"], bold_first=True)
                    r += 1
                    _style_row(ws_analysis, r, 2, ['获得附加分人数', f"{jrb['has_bonus_count']} 人 ({jrb['has_bonus_rate']}%)"], bold_first=True)
                    r += 1
                    # 附加分分布
                    _style_header(ws_analysis, r, 10, ['附加分区间', '人数'])
                    r += 1
                    for label in ['0分', '1-5分', '6-10分', '11-15分', '16-20分']:
                        count = jrb['distribution'].get(label, 0)
                        _style_row(ws_analysis, r, 2, [label, str(count)], bold_first=True)
                        r += 1
                    r += 1

                # ---- 其他项目分析区块 ----
                if other_items:
                    ws_analysis.cell(row=r, column=1, value='测试项目分析').font = subtitle_font
                    r += 1
                    _style_header(ws_analysis, r, 10,
                        ['项目', '优秀', '良好', '及格', '不及格', '有效总数', '优良人数', '优良率'])
                    r += 1
                    for item in other_items:
                        dist = item['distribution']
                        eff = item['effective_total']
                        excellent = dist.get('优秀', 0)
                        good = dist.get('良好', 0)
                        _style_row(ws_analysis, r, 8, [
                            item['item_name'],
                            str(excellent),
                            str(good),
                            str(dist.get('及格', 0)),
                            str(dist.get('不及格', 0)),
                            str(eff),
                            str(excellent + good),
                            f"{item.get('excellence_rate', 0)}%"
                        ], bold_first=True)
                        r += 1
            
            # 各年级概况（全校范围）
            if scope == '全校':
                grade_summaries = analysis_data.get('grade_summaries', [])
                if grade_summaries:
                    r += 1
                    ws_analysis.cell(row=r, column=1, value='各年级概况').font = subtitle_font
                    r += 1
                    _style_header(ws_analysis, r, 10, ['年级', '人数', '优良人数', '优良率', '平均分', '肥胖率'])
                    r += 1
                    for gs in grade_summaries:
                        tg_dist = gs.get('total_grade', {}).get('distribution', {})
                        ec = tg_dist.get('优秀', 0) + tg_dist.get('良好', 0)
                        _style_row(ws_analysis, r, 6, [
                            gs.get('grade_name', ''),
                            str(gs.get('total', 0)),
                            str(ec),
                            f"{gs.get('total_grade', {}).get('excellence_rate', 0)}%",
                            str(gs.get('total_grade', {}).get('avg_score', '')),
                            f"{gs.get('total_grade', {}).get('obesity_rate', 0)}%"
                        ])
                        r += 1
            
            # 班级排名 (仅年级/全校)
            rankings = analysis_data.get('class_rankings', [])
            if rankings:
                r += 1
                ws_analysis.cell(row=r, column=1, value='班级排名（按优良率）').font = subtitle_font
                r += 1
                _style_header(ws_analysis, r, 10, ['排名', '班级', '人数', '优良率', '平均分', '肥胖率'])
                r += 1
                for i, cr in enumerate(rankings):
                    _style_row(ws_analysis, r, 6, [
                        str(i + 1), cr['class_name'], str(cr['total']),
                        f"{cr['excellence_rate']}%", str(cr['avg_score']),
                        f"{cr['obesity_rate']}%"
                    ])
                    r += 1
            
            # 调整列宽
            for ci in range(1, 11):
                ws_analysis.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = 14
        
        # ========== Sheet 2: 学生明细 ==========
        ws_detail = wb.create_sheet('学生明细')
        stats = dm.get_statistics(grade=grade, class_id=class_id)
        
        ws_detail.merge_cells('A1:R1')
        ws_detail.cell(row=1, column=1, value=f'诸葛镇中心小学 — 学生体质健康数据 ({title_map[scope]})').font = title_font
        ws_detail.cell(row=1, column=1).alignment = center_align
        
        headers = [
            '班级编号', '学号', '姓名', '学籍号', '性别',
            '身高(cm)', '体重(kg)', 'BMI', 'BMI等级', 'BMI得分',
            '肺活量', '50米跑', '坐位体前屈', '一分钟跳绳', '跳绳附加分', '仰卧起坐',
            '50*8折返跑', '总分', '等级'
        ]
        _style_header(ws_detail, 2, len(headers), headers)
        
        # 按班级编号、学号排序
        sorted_students = sorted(stats['students'], key=lambda s: (str(s.get('class_id', '') or ''), str(s.get('student_number', '') or '')))
        
        row = 3
        for s in sorted_students:
            data_row = [
                s.get('class_id', '') if s.get('class_id') else (class_id or ''),
                s.get('student_number', ''),
                s.get('name', ''),
                s.get('student_code', ''),
                s.get('gender', ''),
                s.get('height', ''),
                s.get('weight', ''),
                s.get('bmi', ''),
                s.get('bmi_grade', ''),
                s.get('bmi_score', ''),
                s.get('tests', {}).get('肺活量', ''),
                s.get('tests', {}).get('50米跑', ''),
                s.get('tests', {}).get('坐位体前屈', ''),
                s.get('tests', {}).get('一分钟跳绳', ''),
                s.get('jump_rope_bonus', ''),
                s.get('tests', {}).get('仰卧起坐', ''),
                _to_time_str(s.get('tests', {}).get('50*8折返跑')),
                s.get('total_score', ''),
                s.get('total_grade', '')
            ]
            for ci, val in enumerate(data_row, 1):
                cell = ws_detail.cell(row=row, column=ci, value=val if val is not None else '')
                cell.border = border
                cell.alignment = center_align
            row += 1
        
        for ci in range(1, len(headers) + 1):
            ws_detail.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = 12
        
        # ========== Sheet 3: 班级分析 (全校/年级导出时) ==========
        if scope in ('全校', '年级') and analysis_data:
            class_analyses = analysis_data.get('class_analyses', [])
            if class_analyses:
                ws_class = wb.create_sheet('班级分析')
                ws_class.merge_cells('A1:J1')
                ws_class.cell(row=1, column=1, value=f'班级分析汇总 ({title_map[scope]})').font = title_font
                
                all_grade_items = set()
                for ca in class_analyses:
                    for item in ca.get('items', []):
                        all_grade_items.add(item['item_name'])
                
                headers_cls = ['班级', '人数', '男', '女', '有效', '优良人数', '优良率']
                for item_name in sorted(all_grade_items):
                    if item_name != 'BMI':
                        headers_cls.append(f'{item_name}优良率')
                headers_cls += ['BMI正常', 'BMI超重', 'BMI低体重', 'BMI肥胖', '肥胖率']
                
                _style_header(ws_class, 2, len(headers_cls), headers_cls)
                
                row = 3
                for ca in class_analyses:
                    tg_dist = ca.get('total_grade', {}).get('distribution', {})
                    excellence_count = tg_dist.get('优秀', 0) + tg_dist.get('良好', 0)
                    vals = [
                        ca['class_name'], str(ca['total']),
                        str(ca.get('male_count', 0)), str(ca.get('female_count', 0)),
                        str(ca['valid_count']),
                        str(excellence_count),
                        f"{ca['total_grade'].get('excellence_rate', 0)}%"
                    ]
                    
                    for item_name in sorted(all_grade_items):
                        if item_name != 'BMI':
                            found = False
                            for item in ca.get('items', []):
                                if item['item_name'] == item_name:
                                    vals.append(f"{item.get('excellence_rate', 0)}%")
                                    found = True
                                    break
                            if not found:
                                vals.append('-')
                    
                    # BMI 数据
                    bmi_item = next((it for it in ca.get('items', []) if it['item_name'] == 'BMI'), None)
                    if bmi_item:
                        dist = bmi_item['distribution']
                        vals += [str(dist.get('正常', 0)), str(dist.get('超重', 0)),
                                str(dist.get('低体重', 0)), str(dist.get('肥胖', 0)),
                                f"{bmi_item.get('obesity_rate', 0)}%"]
                    else:
                        vals += ['0', '0', '0', '0', '0%']
                    
                    _style_row(ws_class, row, len(headers_cls), vals)
                    row += 1
                
                for ci in range(1, len(headers_cls) + 1):
                    ws_class.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = 13
        
        wb.save(output_path)
        wb.close()
        return True, f"报告已导出到: {output_path}"
    
    except Exception as e:
        logger.exception("导出统计报告失败: %s", output_path)
        return False, f"导出失败: {str(e)}"
