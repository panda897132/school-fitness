"""体测项目数据合成 — 7个分项文件 → 1个完整表格"""

import os
import sys
import tkinter as tk
from tkinter import filedialog, messagebox
import openpyxl
from copy import copy

COL_CLASS = 1
COL_NO = 2
COL_NAME = 3
COL_IDNUM = 4
COL_GENDER = 5
COL_HEIGHT = 6
COL_WEIGHT = 7
COL_VC = 11
COL_RUN50 = 14
COL_SIT_REACH = 17
COL_ROPE = 20
COL_SITUP = 23
COL_SHUTTLE = 26
COL_JUMP_BONUS = 29

COL_MAP = {
    COL_HEIGHT: 6, COL_WEIGHT: 7,
    COL_VC: 11, COL_RUN50: 14,
    COL_SIT_REACH: 17, COL_ROPE: 20,
    COL_SITUP: 23, COL_SHUTTLE: 26,
    COL_JUMP_BONUS: 29,
}

GRADES = ['一年级', '二年级', '三年级', '四年级', '五年级', '六年级']


def merge_all(files):
    if not files:
        return None, '未选择任何文件'
    files = [f for f in files if not os.path.basename(f).startswith('合成')]

    merged = {}
    sheets_map = {g: {} for g in GRADES}

    for fpath in files:
        fname = os.path.basename(fpath)
        if not fname.endswith('.xlsx'):
            continue
        wb = openpyxl.load_workbook(fpath, data_only=True)
        for gs in GRADES:
            if gs not in wb.sheetnames:
                continue
            ws = wb[gs]
            for row in range(2, ws.max_row + 1):
                cid_val = ws.cell(row=row, column=COL_CLASS).value
                name_val = ws.cell(row=row, column=COL_NAME).value
                if not cid_val or not name_val:
                    continue
                cid = str(int(float(str(cid_val))))
                sno = ws.cell(row=row, column=COL_NO).value
                sno = str(int(float(str(sno)))) if sno and '.' in str(sno) else str(sno or '').strip()
                name = str(name_val).strip()
                key = f"{gs}:{cid}:{sno}:{name}"

                if key not in sheets_map[gs]:
                    row_data = {
                        'class_id': cid,
                        'student_no': sno,
                        'name': name,
                        'idnum': '',
                        'gender': '',
                        'cells': {},
                    }
                    for col in (COL_IDNUM, COL_GENDER):
                        v = ws.cell(row=row, column=col).value
                        if col == COL_GENDER:
                            v_raw = str(v or '').strip()
                            v_num = 1 if v_raw in ('1', '男', '') else (2 if v_raw in ('2', '女') else 1)
                            row_data['gender'] = v_num
                        elif col == COL_IDNUM:
                            row_data['idnum'] = str(v or '').strip()
                    sheets_map[gs][key] = row_data
                else:
                    row_data = sheets_map[gs][key]

                for col in COL_MAP:
                    val = ws.cell(row=row, column=col).value
                    if val is None:
                        continue
                    s = str(val).strip()
                    if s in ('', '-', 'None', '#N/A'):
                        continue
                    try:
                        v = float(s) if '.' in s or s.lstrip('-').isdigit() else float(val)
                        sheets_map[gs][key]['cells'][col] = v
                    except (ValueError, TypeError):
                        continue

        wb.close()

    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)

    for gs in GRADES:
        if gs not in sheets_map or not sheets_map[gs]:
            continue
        ws = wb_out.create_sheet(gs)

        header_row = [
            '班级编号', '学号', '姓名', '学籍号', '性别',
            '身高', '体重', 'BMI', 'BMI等级', 'BMI得分',
            '肺活量', '肺活量得分', '肺活量等级',
            '50米跑', '50米跑得分', '50米跑等级',
            '坐位体前屈', '坐位体前屈得分', '坐位体前屈等级',
            '一分钟跳绳', '一分钟跳绳得分', '一分钟跳绳等级',
            '一分钟仰卧起坐', '仰卧起坐得分', '仰卧起坐等级',
            '50米×8往返跑', '往返跑得分', '往返跑等级',
            '跳绳附加分', '标准分', '总分', '总等级'
        ]
        for c, h in enumerate(header_row, 1):
            ws.cell(row=1, column=c, value=h)

        row_idx = 2
        for key, data in sorted(sheets_map[gs].items()):
            ws.cell(row=row_idx, column=1, value=data['class_id'])
            ws.cell(row=row_idx, column=2, value=data['student_no'])
            ws.cell(row=row_idx, column=3, value=data['name'])
            ws.cell(row=row_idx, column=4, value=data['idnum'])
            ws.cell(row=row_idx, column=5, value=data['gender'])
            for col, val in data['cells'].items():
                ws.cell(row=row_idx, column=col, value=val)
            row_idx += 1

    return wb_out, None


def import_merged_excel(filepath, dm):
    """读取合成Excel并导入到系统中（含自动算分）"""
    from score_engine import apply_scores_to_student

    wb = openpyxl.load_workbook(filepath, data_only=True)
    GRADES = ['一年级', '二年级', '三年级', '四年级', '五年级', '六年级']
    COLS = {6: 'height', 7: 'weight', 11: '肺活量', 14: '50米跑',
            17: '坐位体前屈', 20: '一分钟跳绳', 23: '仰卧起坐', 26: '50*8折返跑'}

    total = 0
    for gi, gs in enumerate(GRADES):
        if gs not in wb.sheetnames:
            continue
        ws = wb[gs]
        grade = gi + 1
        class_students = {}

        for row in range(2, ws.max_row + 1):
            cid_val = ws.cell(row=row, column=1).value
            sno_val = ws.cell(row=row, column=2).value
            name_val = ws.cell(row=row, column=3).value
            gender_val = ws.cell(row=row, column=5).value

            if not cid_val or not name_val:
                continue

            cid = str(int(float(str(cid_val))))
            sno = str(int(float(str(sno_val)))) if sno_val and '.' in str(sno_val) else str(sno_val or '').strip()
            g = int(float(str(gender_val))) if gender_val else 1
            gender = '男' if g == 1 else '女'

            sd = {
                'id': f'{cid}{sno}',
                'student_number': sno,
                'name': str(name_val).strip(),
                'gender': gender,
                'height': None, 'weight': None,
                'tests': {},
            }

            for col, item_name in COLS.items():
                val = ws.cell(row=row, column=col).value
                if val is None:
                    continue
                try:
                    v = float(str(val).strip())
                except (ValueError, TypeError):
                    continue
                if col == 6:
                    sd['height'] = v
                elif col == 7:
                    sd['weight'] = v
                else:
                    sd['tests'][item_name] = v

            class_students.setdefault(cid, []).append(sd)

        for cid, students in class_students.items():
            for s in students:
                apply_scores_to_student(s, grade)
            dm.import_students(cid, students)
            total += len(students)

    wb.close()
    dm.flush()
    stats = dm.get_statistics()
    return stats['total'], stats['优秀率'] + stats['良好率'], stats['avg_score']


def main():
    root = tk.Tk()
    root.withdraw()
    filepaths = filedialog.askopenfilenames(
        title='选择体测项目文件（可多选）',
        initialdir=os.path.expanduser('~/桌面/体测项目'),
        filetypes=[('Excel文件', '*.xlsx'), ('所有文件', '*.*')]
    )
    if not filepaths:
        return

    wb, error = merge_all(list(filepaths))
    if error:
        messagebox.showerror('错误', error)
        return

    folder = os.path.dirname(filepaths[0])
    out_path = os.path.join(folder, '合成数据.xlsx')
    wb.save(out_path)
    messagebox.showinfo('完成', f'✅ 合成完成！\n\n已保存: {out_path}\n\n接下来请用"批量导入Excel"导入此文件。')


if __name__ == '__main__':
    main()
